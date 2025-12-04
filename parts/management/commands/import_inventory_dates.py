import re
import unicodedata
from collections import Counter
from datetime import datetime, date, time
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from openpyxl import load_workbook

from parts.inventory_models import PiezaInventario


class Command(BaseCommand):
    """
    Usa la planilla "REPUESTOS STOCK (1).xlsx" para completar la fecha de ingreso
    en registros de PiezaInventario (campo fecha_ingesta_excel).

    La planilla contiene una pestaña llamada "INVENTARIOS" con el par (repuesto, fecha).
    El comando normaliza esos nombres, busca coincidencias en nombre_normalizado / nombre_original
    y actualiza la fecha cuando corresponde.
    """

    help = "Importa fechas de ingreso desde REPUESTOS STOCK (1).xlsx hacia PiezaInventario"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="REPUESTOS STOCK (1).xlsx",
            help="Ruta al Excel de stock manual (por defecto: %(default)s)",
        )
        parser.add_argument(
            "--sheet",
            default="INVENTARIOS",
            help="Nombre de la pestaña con los datos de fechas (por defecto: %(default)s)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Calcula los cambios pero no guarda en la base de datos",
        )
        parser.add_argument(
            "--overwrite",
            action="store_true",
            help="Reemplaza fechas existentes (por defecto solo rellena valores nulos)",
        )

    # --- Helpers ---------------------------------------------------------
    @staticmethod
    def _normalize(value):
        """Normaliza cadenas para facilitar el matching."""
        if not value:
            return ""
        text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
        text = text.lower()
        text = re.sub(r"[^a-z0-9 ]+", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    @staticmethod
    def _parse_date(raw):
        """Acepta datetime, date o strings con formato dd/mm/aaaa."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        if isinstance(raw, str):
            match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", raw)
            if not match:
                return None
            day, month, year = match.groups()
            if len(year) == 2:
                year = f"20{year}"
            try:
                return date(int(year), int(month), int(day))
            except ValueError:
                return None
        return None

    @staticmethod
    def _aware_from_date(date_value):
        """Convierte una fecha (naive) a datetime consciente de zona horaria."""
        naive_dt = datetime.combine(date_value, time.min)
        if timezone.is_naive(naive_dt):
            return timezone.make_aware(naive_dt, timezone.get_current_timezone())
        return naive_dt

    def _build_lookup(self, workbook, sheet_name):
        """
        Construye un diccionario normalizado -> fecha a partir de la pestaña de inventario.
        Se toma la primera coincidencia; el Excel no debería tener duplicados relevantes.
        """
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"La pestaña '{sheet_name}' no existe en el Excel.")

        sheet = workbook[sheet_name]
        lookup = {}
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            _, raw_name, raw_date, *_ = (list(row) + [None, None, None])[:4]
            key = self._normalize(raw_name)
            parsed_date = self._parse_date(raw_date)
            if not key or not parsed_date:
                skipped += 1
                continue
            lookup.setdefault(key, parsed_date)

        self.stdout.write(
            self.style.SUCCESS(
                f"Construido índice de {len(lookup)} repuestos con fecha (omitidos {skipped})."
            )
        )
        return lookup

    # --- Command ---------------------------------------------------------
    def handle(self, *args, **options):
        excel_path = Path(options["file"])
        dry_run = options["dry_run"]
        overwrite = options["overwrite"]
        sheet_name = options["sheet"]

        if not excel_path.exists():
            raise CommandError(f"El archivo '{excel_path}' no existe.")

        self.stdout.write(f"Leyendo Excel: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        lookup = self._build_lookup(workbook, sheet_name)

        if not lookup:
            raise CommandError("No se encontraron filas con nombre+fecha en la pestaña indicada.")

        stats = Counter()
        unmatched_names = set()

        queryset = PiezaInventario.objects.all().only(
            "id", "nombre_normalizado", "nombre_original", "fecha_ingesta_excel"
        )

        for pieza in queryset.iterator(chunk_size=500):
            key = self._normalize(pieza.nombre_normalizado) or self._normalize(pieza.nombre_original)
            if not key:
                stats["sin_nombre"] += 1
                continue

            fecha_excel = lookup.get(key)
            if not fecha_excel:
                stats["sin_fecha_en_excel"] += 1
                unmatched_names.add(key)
                continue

            if pieza.fecha_ingesta_excel and not overwrite:
                stats["conservar_existente"] += 1
                continue

            aware_dt = self._aware_from_date(fecha_excel)
            if pieza.fecha_ingesta_excel == aware_dt:
                stats["sin_cambios"] += 1
                continue

            pieza.fecha_ingesta_excel = aware_dt
            if not dry_run:
                pieza.save(update_fields=["fecha_ingesta_excel"])
            stats["actualizados"] += 1

        summary = ", ".join(f"{k}: {v}" for k, v in sorted(stats.items()))
        if not summary:
            summary = "Sin cambios"

        if dry_run:
            self.stdout.write(self.style.WARNING(f"[DRY-RUN] {summary}"))
        else:
            self.stdout.write(self.style.SUCCESS(summary))

        if unmatched_names:
            ejemplos = ", ".join(sorted(list(unmatched_names))[:10])
            self.stdout.write(
                self.style.HTTP_INFO(
                    f"Sin coincidencia para {len(unmatched_names)} nombres (ej: {ejemplos})"
                )
            )
